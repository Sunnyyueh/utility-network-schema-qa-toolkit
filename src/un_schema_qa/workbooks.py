"""Excel workbook inspection and logical sheet detection."""

from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path

from openpyxl import load_workbook

from .exceptions import InputFormatError, WorkbookDetectionError
from .models.common import StrictModel
from .normalization import resolve_columns


class WorkbookSheet(StrictModel):
    name: str
    headers: tuple[str, ...]
    row_count: int


class WorkbookInspection(StrictModel):
    path: Path
    sheets: tuple[WorkbookSheet, ...]


def inspect_workbook(path: Path | str) -> WorkbookInspection:
    resolved = Path(path).expanduser().resolve()
    try:
        workbook = load_workbook(resolved, read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise InputFormatError(f"cannot open Excel workbook {resolved}: {error}") from error
    sheets: list[WorkbookSheet] = []
    try:
        for worksheet in workbook.worksheets:
            headers: tuple[str, ...] = ()
            row_count = 0
            for values in worksheet.iter_rows(values_only=True):
                if not any(value not in (None, "") for value in values):
                    continue
                if not headers:
                    headers = tuple(
                        str(value).strip() for value in values if value not in (None, "")
                    )
                    continue
                row_count += 1
            sheets.append(WorkbookSheet(name=worksheet.title, headers=headers, row_count=row_count))
    finally:
        workbook.close()
    return WorkbookInspection(path=resolved, sheets=tuple(sheets))


def detect_sheet(inspection: WorkbookInspection, logical_schema: Mapping[str, set[str]]) -> str:
    """Select the single sheet with the strongest logical-column match."""

    scored: list[tuple[int, str]] = []
    for sheet in inspection.sheets:
        if not sheet.headers:
            continue
        score = len(resolve_columns(sheet.headers, logical_schema))
        if score:
            scored.append((score, sheet.name))
    if not scored:
        raise WorkbookDetectionError(f"no matching sheet found in {inspection.path}")
    best_score = max(score for score, _ in scored)
    candidates = sorted(name for score, name in scored if score == best_score)
    if len(candidates) != 1:
        raise WorkbookDetectionError(
            f"ambiguous workbook sheets in {inspection.path}: {', '.join(candidates)}"
        )
    return candidates[0]
