"""Read Excel worksheets into the shared tabular contract."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from un_schema_qa.exceptions import InputFormatError
from un_schema_qa.models import SourceLocation

from .base import TabularRows


def read_xlsx_rows(path: Path, *, sheet: str | None = None) -> TabularRows:
    resolved = path.expanduser().resolve()
    try:
        workbook = load_workbook(resolved, read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise InputFormatError(f"cannot open Excel workbook {resolved}: {error}") from error
    try:
        sheet_name = sheet or _first_nonempty_sheet(workbook)
        if sheet_name not in workbook.sheetnames:
            raise InputFormatError(f"sheet {sheet_name!r} does not exist in {resolved}")
        worksheet = workbook[sheet_name]
        header_row: int | None = None
        headers: tuple[str, ...] = ()
        rows: list[dict[str, Any]] = []
        locations: list[SourceLocation] = []
        for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if not any(value not in (None, "") for value in values):
                continue
            if header_row is None:
                header_row = row_number
                headers = tuple(
                    str(value).strip() if value not in (None, "") else "" for value in values
                )
                if not all(headers):
                    raise InputFormatError(
                        f"blank column name in sheet {sheet_name!r} row {row_number} of {resolved}"
                    )
                if len(headers) != len(set(headers)):
                    raise InputFormatError(
                        f"duplicate column name in sheet {sheet_name!r} of {resolved}"
                    )
                continue
            padded = tuple(values) + (None,) * max(0, len(headers) - len(values))
            row = {header: padded[index] for index, header in enumerate(headers)}
            rows.append(row)
            locations.append(SourceLocation(path=str(resolved), sheet=sheet_name, row=row_number))
    finally:
        workbook.close()
    if not headers or not rows:
        raise InputFormatError(f"empty Excel sheet {sheet_name!r}: {resolved}")
    return TabularRows(headers=headers, rows=tuple(rows), locations=tuple(locations))


def _first_nonempty_sheet(workbook: Workbook) -> str:
    for worksheet in workbook.worksheets:
        for values in worksheet.iter_rows(values_only=True):
            if any(value not in (None, "") for value in values):
                return worksheet.title
    if workbook.active is None:
        raise InputFormatError("Excel workbook has no worksheets")
    return workbook.active.title
